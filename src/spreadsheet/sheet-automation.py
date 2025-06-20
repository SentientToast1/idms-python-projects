import openpyxl as opx
import os
import src.paths as paths


wb = opx.load_workbook(paths.SPREADSHEET)

sheet = wb.active

def searchFunc(name):
    for i in range(2, sheet.max_row + 1):
        if sheet['A'+str(i)].value == name:
            return str(i)
    print("name doesnt exist")
    return None
    


def invokeSearch():
    while True:
        name = input("Enter Name: ")
        searched = searchFunc(name)
        if searched != None:
            return searched
        else:
            option = input("do u want to exit y/n: ")
            if option == 'y' or option == 'Y':
                return None

def dispActive(cell):
    if cell != None:
        print(cell , end='. ')
        print(sheet['A'+cell].value ,end = ' | ')
        print(sheet['B'+cell].value,end = ' | ')
        print(sheet['C'+cell].value,end = ' | ')
        print(sheet['D'+cell].value)
    else:
        print("object does not exist")

def addEntry():
    name = input("Enter name: ")
    joinDate = input("Enter join date in DD-MM-YY: ")
    duration = input("Duration of internship: ")
    project = input("Enter project being worked on: ")
    row = str(sheet.max_row)
    sheet['A' + row] = name
    sheet['B' + row] = joinDate
    sheet['C' + row] = duration
    sheet['D' + row] = project
    wb.save(paths.SPREADSHEET)

def modEntry():
    data = ""


    while True:
        modRow = input("Enter row to be modified: ")
        dispActive(modRow)
        print("Which data do u want to modify? (name, join date, duration, projects) or type exit if u picked wrong option")
        dataId = input("Enter: ").lower()
        if dataId == 'name':
            data = input("Enter new name: ")
            sheet['A' + modRow] = data

        elif dataId == 'join date':
            data = input("Enter new join date: ")
            sheet['B' + modRow] = data

        elif dataId == 'duration':
            data = input("Enter new duration: ")
            sheet['C' + modRow] = data

        elif dataId == 'projects':
            data = input("Enter new project: ")
            sheet['D' + modRow] = data
        elif dataId == 'exit':
            continue
        else:
            print("Wrong input")
        exit = input("modify more data y/n: ")
        if exit == 'y' or exit == 'Y':
            continue
        else:
            break
    wb.save(paths.SPREADSHEET)

def main():
    while True:
        print("Menu:-\n1. Search Data\n2. Add Entry\n3. Delete Entry\n4. Modify Entry\n5. Exit")
        option =input("option: ")
        if option == '1':
            dispActive(invokeSearch())
        elif option == '2':
            addEntry()
        elif option == '3':
            sheet.delete_rows(int(input("Enter row to delete: ")))
            wb.save('intern_list.xlsx')
        elif option == '4':
            modEntry()
        elif option == '5':
            break
        else:
            print("Wrong option")
        
        input("\nPress Enter to continue...")
        os.system('cls')




main()